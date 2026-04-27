from __future__ import annotations

import csv
import re
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Iterable

try:
    import pdfplumber
except ModuleNotFoundError as e:  # pragma: no cover
    raise ModuleNotFoundError(
        'Dependencia ausente: "pdfplumber". Instale com: pip install pdfplumber'
    ) from e


@dataclass(frozen=True)
class TransactionRow:
    data: str
    descricao: str
    valor: str


@dataclass(frozen=True)
class NubankExtractionResult:
    transacoes: list[TransactionRow]
    pagamentos_e_financiamentos: list[TransactionRow]


class NubankTransactionsExtractor:
    """
    Extrai o bloco "TRANSAÇÕES ..." do titular.
    Retorna linhas normalizadas com: Data (YYYY-MM-DD), Descrição, Valor (string "R$ ...").
    """

    MONTHS = {
        "JAN": 1,
        "FEV": 2,
        "MAR": 3,
        "ABR": 4,
        "MAI": 5,
        "JUN": 6,
        "JUL": 7,
        "AGO": 8,
        "SET": 9,
        "OUT": 10,
        "NOV": 11,
        "DEZ": 12,
    }

    DATE_RE = re.compile(r"^(?P<dd>\d{2})\s+(?P<mon>[A-Z]{3})\b")
    # captura "R$ 1.234,56" e também "?R$ 3,94" / "-R$ 3,94"
    VALUE_RE = re.compile(
        r"(?P<sign>[?-]?)\s*R\$\s*(?P<num>\d{1,3}(?:\.\d{3})*,\d{2}|\d+,\d{2})"
    )
    ONLY_BRL_LINE_RE = re.compile(
        r"^(?P<sign>[?-]?)\s*R\$\s*(?P<num>\d{1,3}(?:\.\d{3})*,\d{2}|\d+,\d{2})\s*$"
    )

    def __init__(self, statement_year: int = 2026, holder_name: str | None = "Nome do Titular"):
        self.statement_year = statement_year
        self.holder_name = holder_name

    def extract(self, pdf_path: Path) -> NubankExtractionResult:
        in_transactions = False
        in_holder_block = False
        in_payments = False

        current: TransactionRow | None = None
        transacoes: list[TransactionRow] = []
        pagamentos: list[TransactionRow] = []

        def flush():
            nonlocal current
            if current:
                if in_payments:
                    pagamentos.append(current)
                else:
                    transacoes.append(current)
                current = None

        with pdfplumber.open(str(pdf_path)) as pdf:
            for page in pdf.pages:
                text = page.extract_text() or ""
                for raw_line in text.splitlines():
                    line = raw_line.strip()
                    if not line:
                        continue

                    # O cabeçalho "TRANSAÇÕES DE ..." se repete a cada página. Se a gente
                    # zerar o bloco do titular aqui, só extrai a primeira página.
                    if line.startswith("TRANSAÇÕES DE "):
                        in_transactions = True
                        # Quando holder_name não é informado, extraímos todas as transações do bloco.
                        if self.holder_name is None:
                            in_holder_block = True
                        flush()
                        continue

                    if not in_transactions:
                        continue

                    # Enquanto não entramos em "Pagamentos e Financiamentos", só começamos a
                    # capturar transações após a linha do titular. Depois que entramos em
                    # "Pagamentos e Financiamentos", capturamos independentemente do titular.
                    if (
                        self.holder_name
                        and (not in_payments)
                        and line.startswith(self.holder_name)
                    ):
                        in_holder_block = True
                        flush()
                        continue
                    if self.holder_name and in_payments and line.startswith(self.holder_name):
                        # cabeçalho no topo da página: ignorar para não contaminar descrições
                        continue

                    # entrada do bloco "Pagamentos e Financiamentos"
                    if line.startswith("Pagamentos e Financiamentos"):
                        flush()
                        in_payments = True
                        in_holder_block = False
                        continue

                    # Importante: esta checagem vem DEPOIS do gatilho do titular acima,
                    # senão o extractor nunca "entra" no bloco quando --owner é usado.
                    if self._is_holder_totals_header(line):
                        flush()
                        continue
                    if not in_payments and not in_holder_block:
                        continue

                    # ruídos comuns
                    if re.match(r"^\d+\s*de\s*\d+$", line):
                        continue
                    if "FATURA" in line and "EMISSÃO" in line:
                        continue

                    mdate = self.DATE_RE.match(line)
                    if mdate:
                        flush()
                        dd = int(mdate.group("dd"))
                        mon_abbr = mdate.group("mon")
                        if mon_abbr not in self.MONTHS:
                            continue

                        yyyy = self._parse_year(mon_abbr)
                        mm = self.MONTHS[mon_abbr]
                        d = date(yyyy, mm, dd).isoformat()

                        desc = self._clean_desc(line[mdate.end() :])
                        desc = self._strip_trailing_brl(desc)

                        # Em muitos lançamentos de "Pagamentos e Financiamentos", o PDF já traz o valor final na
                        # mesma linha do favorecido (ex.: "08 MAR Favorecido X ... R$ 10,01"), enquanto o texto
                        # "Total a pagar: R$ 10,00" pode ficar 1 centavo abaixo. Priorizamos o valor da linha da data.
                        if in_payments:
                            valor = self._pick_brl_value_from_line(line)
                        else:
                            valor = self._pick_brl_value_from_line(line)

                        current = TransactionRow(data=d, descricao=desc, valor=valor)
                        continue

                    # continuação da descrição e/ou valor
                    if current:
                        vals = list(self.VALUE_RE.finditer(line))
                        if vals:
                            picked = self._pick_brl_value_from_line(line)
                            if not picked:
                                continue
                            if current.valor and self._is_holder_totals_header(line):
                                continue
                            # Em "Pagamentos e Financiamentos", as linhas com IOF/juros NÃO devem virar valor.
                            # Só aceitamos:
                            # - linha "Total a pagar: ..."
                            # - linha isolada "R$ ..."
                            upper = line.upper().strip()
                            if in_payments:
                                if "TOTAL A PAGAR" in upper:
                                    cur_amt = self._parse_brl_string_to_float(current.valor)
                                    new_amt = self._parse_brl_string_to_float(picked)
                                    if (
                                        cur_amt is not None
                                        and new_amt is not None
                                        and abs(cur_amt - new_amt) <= 0.01001
                                    ):
                                        continue
                                    current = TransactionRow(
                                        data=current.data,
                                        descricao=current.descricao,
                                        valor=picked,
                                    )
                                    continue
                                # linha "R$ 1.711,97" (somente o valor) -> usar.
                                # linha "R$ 2,18 de juros" NÃO deve sobrescrever.
                                if self.ONLY_BRL_LINE_RE.match(line.strip()):
                                    # O PDF costuma trazer duas leituras do mesmo lançamento:
                                    # - "Total a pagar: R$ 51,76" (texto do resumo)
                                    # - linha seguinte só "R$ 51,77" (valor final exibido)
                                    # Quando a diferença é de até 1 centavo, o valor final exibido na fatura é o da
                                    # linha isolada ? usamos esse (e não ficamos presos ao texto do resumo).
                                    cur_amt = self._parse_brl_string_to_float(current.valor)
                                    new_amt = self._parse_brl_string_to_float(picked)
                                    if (
                                        cur_amt is not None
                                        and new_amt is not None
                                        and abs(cur_amt - new_amt) <= 0.01001
                                    ):
                                        current = TransactionRow(
                                            data=current.data,
                                            descricao=current.descricao,
                                            valor=picked,
                                        )
                                        continue
                                    current = TransactionRow(
                                        data=current.data,
                                        descricao=current.descricao,
                                        valor=picked,
                                    )
                                    continue
                                # qualquer outra linha com valores (IOF/juros) é ignorada
                                continue
                            current = TransactionRow(
                                data=current.data,
                                descricao=current.descricao,
                                valor=picked,
                            )
                        else:
                            extra = self._clean_desc(line)
                            if extra:
                                current = TransactionRow(
                                    data=current.data,
                                    descricao=(current.descricao + " " + extra).strip(),
                                    valor=current.valor,
                                )

                # O PDF repete no topo da página o nome completo do titular e outros
                # dados. Sem este flush, essa linha pode ser concatenada na última
                # transação da página anterior.
                flush()

        flush()
        return NubankExtractionResult(
            transacoes=transacoes,
            pagamentos_e_financiamentos=pagamentos,
        )

    def _parse_year(self, mon_abbr: str) -> int:
        # fatura típica: "31 DEZ a 31 JAN" com vencimento/ano 2026 -> DEZ é 2025
        return self.statement_year - 1 if mon_abbr == "DEZ" else self.statement_year

    @staticmethod
    def _brl_to_str(sign: str, num: str) -> str:
        sign = "-" if sign in ("-", "?") else ""
        return f"{sign}R$ {num}"

    @staticmethod
    def _parse_brl_string_to_float(valor: str) -> float | None:
        """Converte 'R$ 1.234,56' / '-R$ 3,94' em float; vazio/ inválido -> None."""
        if not valor or not str(valor).strip():
            return None
        s = str(valor).strip()
        sign = -1.0 if s.startswith("-") or s.startswith("?") else 1.0
        s = s.lstrip("?-").strip()
        s = s.replace("R$", "").strip()
        s = s.replace(".", "").replace(",", ".")
        try:
            return sign * float(s)
        except ValueError:
            return None

    @staticmethod
    def _clean_desc(s: str) -> str:
        s = s.strip()
        # remove máscara do cartão "???? 1234" etc
        s = re.sub(r"^?{4}\s+\d{4}\s+", "", s)
        return re.sub(r"\s+", " ", s).strip()

    def _pick_brl_value_from_line(self, line: str) -> str:
        """
        Nubank pode trazer compras internacionais em múltiplas linhas:
        - "USD 20.00"
        - "Conversão: USD 1 = R$ 5,39"
        - "R$ 99,99"

        Em alguns layouts, a linha de conversão pode conter mais de um "R$".
        Para reduzir falsos positivos no XLSX, quando a linha indica conversão/US$,
        escolhemos o MAIOR valor em BRL encontrado na linha (normalmente o total),
        ao invés do último match (que pode ser a taxa "R$ 5,39").
        """
        vals = list(self.VALUE_RE.finditer(line))
        if not vals:
            return ""

        upper = line.upper()

        # A linha costuma trazer vários "R$" (total, parcelas, IOF, juros). O total cobrado é o primeiro
        # valor imediatamente após "Total a pagar:" (não o maior nem o último da linha inteira).
        if "TOTAL A PAGAR" in upper:
            idx = upper.find("TOTAL A PAGAR")
            tail = line[idx:] if idx >= 0 else line
            m0 = self.VALUE_RE.search(tail)
            if m0:
                return self._brl_to_str(m0.group("sign"), m0.group("num"))
            best = max(vals, key=lambda m: self._brl_match_to_abs_number(m.group("num")))
            return self._brl_to_str(best.group("sign"), best.group("num"))

        # A linha "Conversão: USD 1 = R$ 5,39" contém apenas a taxa e NÃO deve virar o valor da compra.
        # Deixe em branco para não sobrescrever o valor final (que vem em outra linha "R$ 99,99").
        if "CONVERS" in upper and ("USD" in upper or "US$" in upper):
            # Alguns PDFs vêm com taxa + valor final na MESMA linha (ex.: "... = R$ 5,39  R$ 99,99").
            # Se houver 2+ valores em BRL, pegamos o MAIOR (normalmente o total). Se houver só 1, é taxa.
            if len(vals) >= 2:
                best = max(vals, key=lambda m: self._brl_match_to_abs_number(m.group("num")))
                return self._brl_to_str(best.group("sign"), best.group("num"))
            return ""

        # Linha com moeda estrangeira "USD 20.00" não é valor em BRL.
        if upper.startswith("USD ") or upper.startswith("US$ "):
            return ""

        last = vals[-1]
        return self._brl_to_str(last.group("sign"), last.group("num"))

    def _is_holder_totals_header(self, line: str) -> bool:
        """
        Entre blocos de titulares, o Nubank insere linhas do tipo:
        - "Nome do Titular R$ 1.234,56" (subtotal do titular)
        - "Compras de Outro Titular R$ 9.876,54"
        Essas linhas não são transações e não podem sobrescrever valores.
        """
        s = line.strip()
        if not s:
            return False
        # Se esta linha é exatamente o gatilho do titular (passado via --owner),
        # ela não deve ser descartada ANTES de ativar o bloco.
        if self.holder_name and s.startswith(self.holder_name):
            return False
        up = s.upper()
        if up.startswith("PAGAMENTOS E FINANCIAMENTOS"):
            return False
        if up.startswith("COMPRAS DE "):
            return True
        # se começa com data (linha de transação), não é header
        if self.DATE_RE.match(s):
            return False
        # se tem um valor em BRL e parece uma linha curta sem outros marcadores (ex.: IOF de "..."),
        # trata como header de titular/total.
        if "IOF DE" in up or "ESTORNO" in up or "TRANSAÇÃO" in up:
            return False
        vals = list(self.VALUE_RE.finditer(s))
        if not vals:
            return False
        # heurística: termina com valor e tem muitos caracteres (nome do titular)
        last = vals[-1]
        if last.end() != len(s):
            return False
        # evita capturar descrições comuns de compra; headers costumam ter poucas palavras especiais.
        return len(s) <= 80 and any(ch.isalpha() for ch in s)

    @staticmethod
    def _brl_match_to_abs_number(num: str) -> float:
        # num vem como "1.234,56" ou "5,39"
        s = num.replace(".", "").replace(",", ".")
        try:
            return abs(float(s))
        except ValueError:
            return 0.0

    @classmethod
    def _strip_trailing_brl(cls, desc: str) -> str:
        """
        Remove um sufixo do tipo "R$ 123,45" (com sinal opcional) do fim da string.
        Isso evita duplicar o valor na coluna Descrição.
        """
        matches = list(cls.VALUE_RE.finditer(desc))
        if not matches:
            return desc

        last = matches[-1]
        if last.end() == len(desc):
            return desc[: last.start()].rstrip()
        return desc

class SicoobCardStatementExtractor:
    """
    Extrai um PDF do Sicoob (Extrato de Cartão de Crédito).

    - transacoes: bloco "GASTOS DE <titular> ..."
    - pagamentos_e_financiamentos: bloco "MOVIMENTOS" (saldo anterior, pagamentos, encargos etc.)
    """

    DATE_DDMM_RE = re.compile(r"^(?P<dd>\d{2})/(?P<mm>\d{2})\b")
    DATE_DDMMYYYY_RE = re.compile(r"\b(?P<dd>\d{2})/(?P<mm>\d{2})/(?P<yyyy>\d{4})\b")
    # último número do tipo 1.234,56 ou -30,00 (sem "R$")
    LAST_VALUE_RE = re.compile(r"(?P<num>-?\d{1,3}(?:\.\d{3})*,\d{2})\s*$")

    def __init__(self, statement_year: int = 2026):
        self.statement_year = statement_year

    def extract(self, pdf_path: Path) -> NubankExtractionResult:
        in_movimentos = False
        in_gastos = False

        current: TransactionRow | None = None
        transacoes: list[TransactionRow] = []
        movimentos: list[TransactionRow] = []

        statement_month = 1
        statement_year = self.statement_year
        valid_months_gastos: set[int] = {1, 12}

        def flush():
            nonlocal current
            if current:
                if in_movimentos and not in_gastos:
                    movimentos.append(current)
                elif in_gastos:
                    transacoes.append(current)
                current = None

        with pdfplumber.open(str(pdf_path)) as pdf:
            # tenta inferir mês/ano da fatura a partir do cabeçalho (ex.: 31/01/2026 ...)
            for page in pdf.pages[:1]:
                header_text = page.extract_text() or ""
                m = self.DATE_DDMMYYYY_RE.search(header_text)
                if m:
                    statement_year = int(m.group("yyyy"))
                    statement_month = int(m.group("mm"))
                    prev_month = 12 if statement_month == 1 else (statement_month - 1)
                    valid_months_gastos = {statement_month, prev_month}
                    break

            for page in pdf.pages:
                text = page.extract_text() or ""
                for raw_line in text.splitlines():
                    line = raw_line.strip()
                    if not line:
                        continue

                    if line.startswith("MOVIMENTOS"):
                        flush()
                        in_movimentos = True
                        in_gastos = False
                        continue

                    if line.startswith("GASTOS DE "):
                        flush()
                        in_gastos = True
                        in_movimentos = False
                        continue

                    # fim do bloco de gastos
                    if in_gastos and (line.startswith("TOTAL ") or line.startswith("DEMONSTRATIVO")):
                        flush()
                        in_gastos = False
                        continue

                    if not in_movimentos and not in_gastos:
                        continue

                    # ignora cabeçalhos/ruído
                    if line == "SICOOB" or "EXTRATO DE CARTÃO DE CRÉDITO" in line:
                        continue
                    if line.startswith("Cliente:") or line.startswith("Fatura de "):
                        continue

                    # linha "SALDO ANTERIOR ..." (sem data)
                    if in_movimentos and line.startswith("- SALDO ANTERIOR"):
                        flush()
                        val = self._parse_last_value_as_brl(line)
                        desc = line.lstrip("-").strip()
                        desc = self._strip_trailing_value(desc)
                        current = TransactionRow(
                            data="",
                            descricao=f"SALDO ANTERIOR: {desc.replace('SALDO ANTERIOR', '').strip()}",
                            valor=val or "",
                        )
                        flush()
                        continue

                    # tratar quebra de linha que começa com "01/02" (parcela) etc
                    # No Sicoob, compras parceladas aparecem como "01/02", "02/03" (parcela),
                    # e quando a descrição quebra, isso pode vir no começo da linha, parecendo
                    # uma data DD/MM. Só fundimos nesse caso quando a linha anterior ainda não
                    # tem valor (continuação/parcela). Se o valor já veio na linha anterior,
                    # DD/MM no início é novo lançamento: o bloco GASTOS lista compras de vários
                    # meses, não só o mês da fatura e o anterior.
                    # BEGIN sicoob-gastos-date-vs-parcela
                    if current and in_gastos and not current.valor:
                        mstart = self.DATE_DDMM_RE.match(line)
                        if mstart:
                            mm0 = int(mstart.group("mm"))
                            if mm0 not in valid_months_gastos:
                                current = self._merge_continuation(current, line)
                                continue
                    # END sicoob-gastos-date-vs-parcela

                    mdate = self.DATE_DDMM_RE.match(line)
                    if mdate:
                        flush()
                        dd = int(mdate.group("dd"))
                        mm = int(mdate.group("mm"))
                        yyyy = statement_year - 1 if mm > statement_month else statement_year
                        d = date(yyyy, mm, dd).isoformat()

                        rest = line[mdate.end() :].strip()
                        valor = (
                            self._parse_last_value_as_brl(rest)
                            or self._parse_last_value_as_brl(line)
                            or ""
                        )
                        desc = self._strip_trailing_value(rest)
                        current = TransactionRow(data=d, descricao=desc, valor=valor)
                        continue

                    # continuação (quebras de linha comuns no PDF)
                    if current:
                        current = self._merge_continuation(current, line)

                # evita concatenar cabeçalho da página seguinte
                flush()

        flush()
        return NubankExtractionResult(
            transacoes=transacoes,
            pagamentos_e_financiamentos=movimentos,
        )

    def _parse_last_value_as_brl(self, s: str) -> str | None:
        m = self.LAST_VALUE_RE.search(s.strip())
        if not m:
            return None
        num = m.group("num")
        sign = "-" if num.startswith("-") else ""
        num_clean = num[1:] if sign else num
        return f"{sign}R$ {num_clean}"

    def _strip_trailing_value(self, s: str) -> str:
        m = self.LAST_VALUE_RE.search(s.strip())
        if not m:
            return s.strip()
        return s[: m.start()].rstrip()

    # helper único para merge de continuação
    def _merge_continuation(self, current: TransactionRow, line: str) -> TransactionRow:
        """
        Junta uma linha de continuação no lançamento atual.

        Regras:
        - Se a linha termina com um valor (ex.: "CANED 69,15"), anexa a parte textual
          na descrição e preenche o valor (somente se ainda estiver vazio).
        - Caso contrário, concatena a linha inteira na descrição.
        """
        # evita concatenar linhas de cabeçalho entre páginas
        if line.startswith("Cliente:") or line.startswith("Conta Cartão:"):
            return current

        mval = self.LAST_VALUE_RE.search(line.strip())
        if mval:
            before = line[: mval.start()].strip()
            updated = current
            if before:
                updated = TransactionRow(
                    data=updated.data,
                    descricao=(updated.descricao + " " + before).strip(),
                    valor=updated.valor,
                )
            # Se a linha de continuação termina com valor:
            # - Preenche o valor somente se ainda estiver vazio.
            # - Se já existe valor, evita sobrescrever e só usa o texto "before".
            if not updated.valor:
                updated = TransactionRow(
                    data=updated.data,
                    descricao=updated.descricao,
                    valor=self._parse_last_value_as_brl(line) or "",
                )
            return updated

        return TransactionRow(
            data=current.data,
            descricao=(current.descricao + " " + line).strip(),
            valor=current.valor,
        )


def detect_bank(pdf_path: Path) -> str:
    """
    Identifica o banco do PDF: 'nubank', 'sicoob' ou 'santander'.
    """
    name = pdf_path.name.lower()
    if "sicoob" in name:
        return "sicoob"
    if "nubank" in name:
        return "nubank"
    if "santander" in name:
        return "santander"

    with pdfplumber.open(str(pdf_path)) as pdf:
        first = (pdf.pages[0].extract_text() or "").upper()
    if "SICOOB" in first and "EXTRATO DE CARTÃO DE CRÉDITO" in first:
        return "sicoob"
    if "BANCO SANTANDER" in first or "APP WAY" in first or "DETALHAMENTO DA FATURA" in first:
        return "santander"
    if "NU PAGAMENTOS" in first or "RESUMO DA FATURA ATUAL" in first:
        return "nubank"
    # fallback
    return "nubank"


class SantanderCardStatementExtractor:
    """
    Extrai a seção "Detalhamento da Fatura" do Santander (App Way).

    A fatura traz um bloco com tabelas ("Parcelamentos" / "Despesas") e termina antes de
    "Saldo total consolidado de obrigações futuras". As linhas variam:
    - "3 11/04 MED*FACIL 141,80" (com número da compra)
    - "14/04 PGTO BOLETO 430,38" (sem número da compra)
    - "TARIFA PAGAMENTOCONTAS 15,02" (sem data; assume a última data vista no bloco)
    """

    DATE_DDMMYYYY_RE = re.compile(r"\b(?P<dd>\d{2})/(?P<mm>\d{2})/(?P<yyyy>\d{4})\b")
    LAST_VALUE_RE = re.compile(r"(?P<num>-?\d{1,3}(?:\.\d{3})*,\d{2})\s*$")

    # linha de compra típica (compra opcional, parcela opcional)
    # exemplos:
    # - "3 11/04 MED*FACIL 141,80"
    # - "2 03/04 MP*DECOMOBILIDF 01/12 499,24"
    # - "14/04 PGTO BOLETO 430,38"
    LINE_RE = re.compile(
        r"^(?:(?P<compra>\d+)\s+)?"
        r"(?P<dd>\d{2})/(?P<mm>\d{2})\s+"
        r"(?P<desc>.+?)\s+"
        r"(?:(?P<parc>\d{2}/\d{2})\s+)?"
        r"(?P<brl>\d{1,3}(?:\.\d{3})*,\d{2})"
        r"(?:\s+(?P<usd>\d{1,3}(?:\.\d{3})*,\d{2}))?\s*$"
    )

    # encontra inícios de registros (inclui "compra" se existir)
    SPLIT_ANCHOR_RE = re.compile(r"(?:(?<=\s)|^)(?:\d+\s+)?\d{2}/\d{2}\b")

    def __init__(self, statement_year: int = 2026):
        self.statement_year = statement_year

    def extract(self, pdf_path: Path) -> NubankExtractionResult:
        transacoes: list[TransactionRow] = []
        outros: list[TransactionRow] = []

        statement_month = 1
        statement_year = self.statement_year
        last_date_iso: str | None = None
        in_detail = False
        current: TransactionRow | None = None

        with pdfplumber.open(str(pdf_path)) as pdf:
            # tenta inferir ano/mês varrendo as primeiras páginas por DD/MM/YYYY (ex.: Vencimento 01/05/2026)
            for page in pdf.pages[:3]:
                header_text = page.extract_text() or ""
                m = self.DATE_DDMMYYYY_RE.search(header_text)
                if m:
                    statement_year = int(m.group("yyyy"))
                    statement_month = int(m.group("mm"))
                    break

            for page in pdf.pages:
                # Em PDFs do Santander, `extract_text()` pode perder linhas inteiras.
                # Usamos `extract_words()` e recortamos a área entre:
                # "Detalhamento da Fatura" e ("VALOR TOTAL"/"Saldo total consolidado...").
                lines = self._extract_detail_lines(page)
                if not lines:
                    continue

                normalized: list[str] = []
                pending_compra: str | None = None
                pending_value: str | None = None
                for ln in lines:
                    s = ln.strip()
                    if not s:
                        continue
                    if re.fullmatch(r"\d{1,3}", s):
                        pending_compra = s
                        continue
                    # algumas linhas vêm como "15,02" (valor) e na linha seguinte a descrição ("TARIFA ...")
                    if re.fullmatch(r"\d{1,3}(?:\.\d{3})*,\d{2}", s):
                        pending_value = s
                        continue
                    if pending_value and not re.match(r"^(?:\d+\s+)?\d{2}/\d{2}\b", s):
                        s = f"{s} {pending_value}"
                        pending_value = None
                    if pending_compra:
                        s = f"{pending_compra} {s}"
                        pending_compra = None
                    normalized.append(s)

                for raw in normalized:
                    upper = raw.upper()

                    # fim do bloco (segurança extra)
                    # A coluna da direita às vezes aparece "no meio" do texto; não pode encerrar parsing.
                    if "SALDO TOTAL CONSOLIDADO DE OBRIGAÇÕES FUTURAS" in upper:
                        continue
                    if "RESUMO DA FATURA" in upper:
                        # às vezes vem colado no fim de uma linha; truncar e continuar
                        raw = raw.split("Resumo da Fatura", 1)[0].strip()
                        upper = raw.upper()
                        if not raw:
                            break

                    # "Despesas  VALOR TOTAL 7.959,14 0,00" é cabeçalho/total, não é fim do bloco.
                    if "VALOR TOTAL" in upper:
                        continue
                    if "COMPRAS PARCELADAS" in upper or "CRÉDITO E TARIFAS" in upper:
                        continue

                    # ruídos do resumo (coluna da direita) e cabeçalhos
                    if upper.startswith("DESCRI") and "R$" in upper:
                        continue
                    if upper.startswith("SALDO ANTERIOR") or upper.startswith("(+)") or upper.startswith("(-)") or upper.startswith("(=)"):
                        continue

                    # ignora cabeçalhos das tabelas
                    if upper in ("PARCELAMENTOS", "DESPESAS"):
                        continue
                    if upper.startswith("COMPRA DATA") and "DESCRI" in upper and "R$" in upper:
                        continue

                    for line in self._split_merged_rows(raw):
                        # normaliza quando o número da compra "vai parar" no fim da linha
                        # ex.: "21/04 FORNECEDORXYZ 3 30,00" -> "3 21/04 FORNECEDORXYZ 30,00"
                        line = re.sub(
                            r"^(?P<dd>\d{2})/(?P<mm>\d{2})\s+(?P<desc>.+?)\s+(?P<compra>\d+)\s+(?P<val>\d{1,3}(?:\.\d{3})*,\d{2})\s*$",
                            r"\g<compra> \g<dd>/\g<mm> \g<desc> \g<val>",
                            line,
                        ).strip()

                        # remove prefixo de cabeçalho do cartão quando ele "cola" com a 1ª compra da linha
                        # ex.: "NOME ... 1234 XXXX XXXX 5678 21/04 LOJA ... 123,45"
                        line = re.sub(
                            r"^.*?\d{4}\s+X{4}\s+X{4}\s+\d{4}\s+",
                            "",
                            line,
                            flags=re.I,
                        ).strip()

                        line = re.sub(r"^(PARCELAMENTOS|DESPESAS)\s+", "", line, flags=re.I).strip()
                        line = re.sub(r"^COMPRA\s+DATA\s+DESCRI.+?\s+", "", line, flags=re.I).strip()
                        line = re.sub(r"^\d+\s+COMPRA\s+DATA\s+DESCRI.+?\s+", "", line, flags=re.I).strip()
                        line = re.sub(r"^PARCELA\s+R\$\s+US\$\s+", "", line, flags=re.I).strip()
                        if not line:
                            continue

                        mrow = self.LINE_RE.match(line)
                        if mrow:
                            if current:
                                transacoes.append(current)
                                current = None

                            dd = int(mrow.group("dd"))
                            mm = int(mrow.group("mm"))
                            yyyy = statement_year - 1 if mm > statement_month else statement_year
                            d_iso = date(yyyy, mm, dd).isoformat()
                            last_date_iso = d_iso

                            desc = mrow.group("desc").strip()
                            parc = (mrow.group("parc") or "").strip()
                            if parc:
                                desc = f"{desc} {parc}".strip()

                            val = f"R$ {mrow.group('brl')}"
                            current = TransactionRow(data=d_iso, descricao=desc, valor=val)
                            continue

                        mval = self.LAST_VALUE_RE.search(line)
                        if mval and last_date_iso:
                            if current:
                                transacoes.append(current)
                                current = None
                            val = self._parse_last_value_as_brl(line) or ""
                            desc = self._strip_trailing_value(line)
                            if desc:
                                current = TransactionRow(data=last_date_iso, descricao=desc, valor=val)
                            continue

                        if current:
                            current = TransactionRow(
                                data=current.data,
                                descricao=(current.descricao + " " + line).strip(),
                                valor=current.valor,
                            )

        if current:
            transacoes.append(current)

        return NubankExtractionResult(transacoes=transacoes, pagamentos_e_financiamentos=outros)

    def _parse_last_value_as_brl(self, s: str) -> str | None:
        m = self.LAST_VALUE_RE.search(s.strip())
        if not m:
            return None
        num = m.group("num")
        sign = "-" if num.startswith("-") else ""
        num_clean = num[1:] if sign else num
        return f"{sign}R$ {num_clean}"

    def _strip_trailing_value(self, s: str) -> str:
        m = self.LAST_VALUE_RE.search(s.strip())
        if not m:
            return s.strip()
        return s[: m.start()].rstrip()

    def _split_merged_rows(self, line: str) -> list[str]:
        """
        O pdfplumber às vezes "cola" duas linhas de compra numa única linha, por exemplo:
        "3 21/04 ... 50,45 3 22/04 ... 18,00"
        Esta função divide a string em partes candidatas para parsing.
        """
        s = line.strip()
        if not s:
            return []

        anchors = list(self.SPLIT_ANCHOR_RE.finditer(s))
        if len(anchors) <= 1:
            return [s]

        # divide em todos os inícios de data (a partir do segundo)
        parts: list[str] = []
        start = 0
        for a in anchors[1:]:
            idx = a.start()
            chunk = s[start:idx].strip()
            if chunk:
                parts.append(chunk)
            start = idx
        tail = s[start:].strip()
        if tail:
            parts.append(tail)
        return parts

    def _extract_detail_lines(self, page) -> list[str]:
        """
        Retorna linhas (texto) apenas do bloco "Detalhamento da Fatura" desta página.
        Se a página não contém o bloco, retorna [].
        """
        words = page.extract_words() or []
        if not words:
            return []

        def find_min_top(needle: str) -> float | None:
            tops = [w["top"] for w in words if needle in w["text"].upper()]
            return min(tops) if tops else None

        detail_top = find_min_top("DETALHAMENTO")
        if detail_top is None:
            return []

        # O layout pode posicionar "VALOR TOTAL" perto do topo (mesma linha do cabeçalho "Despesas"),
        # e tentativas de achar um "fim" por palavras acabam cortando linhas válidas.
        # Por isso, extraímos até o fim da página e paramos por palavras-chave durante o parsing.
        y0 = min(detail_top + 5, page.height)
        y1 = page.height

        cropped = page.within_bbox((0, y0, page.width, y1))
        cwords = cropped.extract_words() or []
        if not cwords:
            return []

        # A página tem layout em colunas. Para não perder palavras (como no split por word),
        # primeiro reconstruímos a linha completa (todas as palavras ordenadas por x0),
        # depois classificamos a linha na coluna esquerda/direita pelo menor x0.
        split_x = page.width * 0.55

        by_top: dict[int, list[dict]] = {}
        for w in cwords:
            top_key = int(round(w["top"]))
            by_top.setdefault(top_key, []).append(w)

        reconstructed: list[tuple[int, int, str]] = []  # (col, top_key, text)
        for top_key in sorted(by_top.keys()):
            row_words = sorted(by_top[top_key], key=lambda ww: ww["x0"])
            text = " ".join(ww["text"] for ww in row_words).strip()
            if not text:
                continue
            if "DETALHAMENTO" in text.upper() and len(text) < 40:
                continue
            min_x0 = min(ww["x0"] for ww in row_words)
            col = 0 if min_x0 < split_x else 1
            reconstructed.append((col, top_key, text))

        reconstructed.sort(key=lambda t: (t[0], t[1]))  # esquerda (0) depois direita (1)
        return [t[2] for t in reconstructed]

class CsvWriter:
    def write_report(self, result: NubankExtractionResult, out_csv_path: Path) -> None:
        out_csv_path.parent.mkdir(parents=True, exist_ok=True)
        with out_csv_path.open("w", newline="", encoding="utf-8") as f:
            w = csv.writer(f)
            w.writerow(["Data", "Descrição", "Valor"])

            for r in result.transacoes:
                w.writerow([r.data, r.descricao, r.valor])

            for r in result.pagamentos_e_financiamentos:
                # mantém o mesmo formato de 3 colunas, mas marca a origem
                w.writerow([r.data, f"[Pagamentos e Financiamentos] {r.descricao}", r.valor])


class XlsxWriter:
    def write_report(self, result: NubankExtractionResult, out_xlsx_path: Path) -> None:
        try:
            from openpyxl import Workbook
        except ModuleNotFoundError as e:  # pragma: no cover
            raise ModuleNotFoundError(
                'Dependência ausente: "openpyxl". Instale com: pip install openpyxl'
            ) from e

        out_xlsx_path.parent.mkdir(parents=True, exist_ok=True)

        # Observação: Excel/locale podem variar, mas este formato costuma funcionar bem:
        # - moeda "R$"
        # - negativos em parênteses (e vermelhos)
        # - traço para zero
        brl_accounting_format = (
            '_("R$"* #,##0.00_);[Red]_("R$"* (#,##0.00);_("R$"* "-"??_);_(@_)'
        )

        wb = Workbook()
        ws_trans = wb.active
        ws_trans.title = "Transações"

        ws_trans.append(["Data", "Descrição", "Valor"])
        for r in result.transacoes:
            self._append_row_with_accounting(ws_trans, r, brl_accounting_format)

        ws_pay = wb.create_sheet("Pagamentos e Financiamentos")
        ws_pay.append(["Data", "Descrição", "Valor"])
        for r in result.pagamentos_e_financiamentos:
            self._append_row_with_accounting(ws_pay, r, brl_accounting_format)

        wb.save(str(out_xlsx_path))

    @staticmethod
    def _append_row_with_accounting(ws, r: TransactionRow, brl_accounting_format: str) -> None:
        """
        Escreve (Data, Descrição, Valor) onde Valor vira número e recebe formato Contabilidade.
        """
        ws.append([r.data, r.descricao, None])
        cell = ws.cell(row=ws.max_row, column=3)
        cell.value = XlsxWriter._parse_brl_to_number(r.valor)
        cell.number_format = brl_accounting_format

    @staticmethod
    def _parse_brl_to_number(valor: str):
        """
        Converte strings como "R$ 1.234,56" / "-R$ 3,94" em número (float) para Excel.
        Se estiver vazio/indefinido, retorna None.
        """
        if not valor:
            return None

        s = valor.strip()
        sign = -1 if s.startswith("-") or s.startswith("?") else 1
        s = s.lstrip("?-").strip()
        s = s.replace("R$", "").strip()

        # remove separador de milhar "." e troca decimal "," -> "."
        s = s.replace(".", "").replace(",", ".")
        try:
            return round(sign * float(s), 2)
        except ValueError:
            return None
